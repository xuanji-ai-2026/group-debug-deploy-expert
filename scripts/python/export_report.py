#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报表导出脚本 - 商机统计与计费报表
================================================
Author: 孙丽 (EMP-DATA-002)
Version: 1.0.0
"""

import os
import sys
import json
import logging
import argparse
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, asdict
from decimal import Decimal

import pymysql
from pymysql.cursors import DictCursor
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(f'export_report_{datetime.now():%Y%m%d}.log')
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class ReportConfig:
    """报表配置"""
    report_type: str  # leads, billing, combined
    start_date: datetime
    end_date: datetime
    output_format: str = 'excel'  # excel, csv, json
    tenant_id: Optional[str] = None
    output_dir: str = './reports'


class ReportExporter:
    """报表导出器"""
    
    def __init__(self, host: str, port: int, user: str, password: str, 
                 database: str, charset: str = 'utf8mb4'):
        self.config = {
            'host': host,
            'port': port,
            'user': user,
            'password': password,
            'database': database,
            'charset': charset,
            'cursorclass': DictCursor
        }
        self.conn = None
    
    def connect(self) -> bool:
        """建立数据库连接"""
        try:
            self.conn = pymysql.connect(**self.config)
            logger.info("数据库连接成功")
            return True
        except Exception as e:
            logger.error(f"数据库连接失败: {e}")
            return False
    
    def close(self):
        """关闭连接"""
        if self.conn:
            self.conn.close()
            logger.info("数据库连接已关闭")
    
    def query_to_dataframe(self, sql: str, params: Tuple = ()) -> pd.DataFrame:
        """执行查询并返回DataFrame"""
        return pd.read_sql(sql, self.conn, params=params)
    
    def generate_lead_report(self, start_date: datetime, 
                            end_date: datetime,
                            tenant_id: Optional[str] = None) -> pd.DataFrame:
        """
        生成商机统计报表
        """
        logger.info(f"生成商机报表: {start_date.date()} ~ {end_date.date()}")
        
        tenant_filter = "AND l.tenant_id = %s" if tenant_id else ""
        params = (start_date, end_date)
        if tenant_id:
            params += (tenant_id,)
        
        sql = f"""
        SELECT 
            l.tenant_id,
            t.name as tenant_name,
            l.status,
            l.source,
            COUNT(*) as lead_count,
            SUM(l.estimated_value) as total_estimated,
            SUM(l.actual_value) as total_actual,
            AVG(l.probability) as avg_probability,
            COUNT(DISTINCT l.owner_id) as active_users,
            MIN(l.created_at) as first_lead_date,
            MAX(l.created_at) as last_lead_date
        FROM leads l
        LEFT JOIN tenants t ON l.tenant_id = t.id
        WHERE l.created_at BETWEEN %s AND %s
        {tenant_filter}
        GROUP BY l.tenant_id, l.status, l.source
        ORDER BY l.tenant_id, l.status
        """
        
        return self.query_to_dataframe(sql, params)
    
    def generate_lead_summary(self, start_date: datetime,
                             end_date: datetime,
                             tenant_id: Optional[str] = None) -> pd.DataFrame:
        """生成商机汇总报表"""
        tenant_filter = "AND tenant_id = %s" if tenant_id else ""
        params = (start_date, end_date)
        if tenant_id:
            params += (tenant_id,)
        
        sql = f"""
        SELECT 
            status,
            COUNT(*) as count,
            SUM(estimated_value) as estimated_total,
            SUM(CASE WHEN status = 'won' THEN actual_value ELSE 0 END) as won_value,
            AVG(probability) as avg_probability
        FROM leads
        WHERE created_at BETWEEN %s AND %s
        {tenant_filter}
        GROUP BY status
        ORDER BY FIELD(status, 'new', 'contacted', 'qualified', 'proposal', 'negotiation', 'won', 'lost')
        """
        
        return self.query_to_dataframe(sql, params)
    
    def generate_lead_funnel(self, start_date: datetime,
                            end_date: datetime,
                            tenant_id: Optional[str] = None) -> pd.DataFrame:
        """生成销售漏斗报表"""
        tenant_filter = "AND tenant_id = %s" if tenant_id else ""
        params = (start_date, end_date)
        if tenant_id:
            params += (tenant_id,)
        
        sql = f"""
        SELECT 
            status,
            COUNT(*) as lead_count,
            ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) as percentage,
            SUM(estimated_value) as pipeline_value,
            AVG(DATEDIFF(COALESCE(closed_at, NOW()), created_at)) as avg_days
        FROM leads
        WHERE created_at BETWEEN %s AND %s
        {tenant_filter}
        GROUP BY status
        ORDER BY FIELD(status, 'new', 'contacted', 'qualified', 'proposal', 'negotiation', 'won', 'lost')
        """
        
        return self.query_to_dataframe(sql, params)
    
    def generate_billing_report(self, start_date: datetime,
                               end_date: datetime,
                               tenant_id: Optional[str] = None) -> pd.DataFrame:
        """
        生成计费统计报表
        """
        logger.info(f"生成计费报表: {start_date.date()} ~ {end_date.date()}")
        
        # 租户订阅统计
        tenant_filter = "WHERE t.id = %s" if tenant_id else ""
        params = (tenant_id,) if tenant_id else ()
        
        sql = f"""
        SELECT 
            t.id as tenant_id,
            t.name as tenant_name,
            t.plan_type,
            t.status as tenant_status,
            COUNT(DISTINCT u.id) as user_count,
            t.max_users,
            t.storage_limit_gb,
            DATEDIFF(t.expire_at, NOW()) as days_to_expire,
            t.created_at as tenant_created_at,
            t.expire_at as subscription_expire_at
        FROM tenants t
        LEFT JOIN users u ON t.id = u.tenant_id AND u.status = 'active'
        {tenant_filter}
        GROUP BY t.id
        ORDER BY t.plan_type, t.created_at
        """
        
        return self.query_to_dataframe(sql, params)
    
    def generate_revenue_report(self, start_date: datetime,
                               end_date: datetime) -> pd.DataFrame:
        """生成收入统计报表"""
        # 按套餐类型统计
        sql = """
        SELECT 
            plan_type,
            COUNT(*) as tenant_count,
            SUM(max_users) as total_licenses,
            SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) as active_count,
            SUM(CASE WHEN status = 'suspended' THEN 1 ELSE 0 END) as suspended_count,
            SUM(CASE WHEN expire_at < DATE_ADD(NOW(), INTERVAL 30 DAY) THEN 1 ELSE 0 END) as expiring_soon
        FROM tenants
        GROUP BY plan_type
        ORDER BY FIELD(plan_type, 'free', 'basic', 'pro', 'enterprise')
        """
        
        return self.query_to_dataframe(sql)
    
    def generate_user_activity_report(self, start_date: datetime,
                                     end_date: datetime,
                                     tenant_id: Optional[str] = None) -> pd.DataFrame:
        """生成用户活跃度报表"""
        tenant_filter = "AND tenant_id = %s" if tenant_id else ""
        params = (start_date, end_date)
        if tenant_id:
            params += (tenant_id,)
        
        sql = f"""
        SELECT 
            tenant_id,
            role,
            COUNT(*) as user_count,
            SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) as active_count,
            SUM(CASE WHEN last_login_at > DATE_SUB(NOW(), INTERVAL 7 DAY) THEN 1 ELSE 0 END) as active_last_7d,
            SUM(CASE WHEN last_login_at > DATE_SUB(NOW(), INTERVAL 30 DAY) THEN 1 ELSE 0 END) as active_last_30d,
            AVG(DATEDIFF(NOW(), created_at)) as avg_account_age
        FROM users
        WHERE created_at BETWEEN %s AND %s
        {tenant_filter}
        GROUP BY tenant_id, role
        ORDER BY tenant_id
        """
        
        return self.query_to_dataframe(sql, params)
    
    def export_to_excel(self, dataframes: Dict[str, pd.DataFrame], 
                        output_path: str):
        """导出到Excel文件"""
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in dataframes.items():
                if not df.empty:
                    # 限制sheet名称长度
                    safe_name = sheet_name[:31]
                    df.to_excel(writer, sheet_name=safe_name, index=False)
                    
                    # 获取工作簿和工作表进行样式设置
                    workbook = writer.book
                    if safe_name in workbook.sheetnames:
                        worksheet = workbook[safe_name]
                        
                        # 设置表头样式
                        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        header_font = Font(color='FFFFFF', bold=True)
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 自动调整列宽
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Excel报表已导出: {output_path}")
    
    def export_to_csv(self, df: pd.DataFrame, output_path: str):
        """导出到CSV文件"""
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        logger.info(f"CSV报表已导出: {output_path}")
    
    def export_to_json(self, data: Dict, output_path: str):
        """导出到JSON文件"""
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        
        def json_serial(obj):
            if isinstance(obj, (datetime, timedelta)):
                return obj.isoformat()
            if isinstance(obj, Decimal):
                return float(obj)
            raise TypeError(f"Type {type(obj)} not serializable")
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=json_serial)
        
        logger.info(f"JSON报表已导出: {output_path}")
    
    def generate_full_report(self, config: ReportConfig) -> str:
        """生成完整报表"""
        os.makedirs(config.output_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        dataframes = {}
        
        if config.report_type in ['leads', 'combined']:
            # 商机报表
            dataframes['商机明细'] = self.generate_lead_report(
                config.start_date, config.end_date, config.tenant_id
            )
            dataframes['商机汇总'] = self.generate_lead_summary(
                config.start_date, config.end_date, config.tenant_id
            )
            dataframes['销售漏斗'] = self.generate_lead_funnel(
                config.start_date, config.end_date, config.tenant_id
            )
        
        if config.report_type in ['billing', 'combined']:
            # 计费报表
            dataframes['租户订阅'] = self.generate_billing_report(
                config.start_date, config.end_date, config.tenant_id
            )
            dataframes['收入统计'] = self.generate_revenue_report(
                config.start_date, config.end_date
            )
        
        if config.report_type == 'combined':
            # 用户活跃度
            dataframes['用户活跃度'] = self.generate_user_activity_report(
                config.start_date, config.end_date, config.tenant_id
            )
        
        # 导出
        if config.output_format == 'excel':
            output_path = os.path.join(config.output_dir, f'report_{timestamp}.xlsx')
            self.export_to_excel(dataframes, output_path)
        
        elif config.output_format == 'csv':
            output_paths = []
            for name, df in dataframes.items():
                safe_name = name.replace(' ', '_')
                output_path = os.path.join(config.output_dir, f'{safe_name}_{timestamp}.csv')
                self.export_to_csv(df, output_path)
                output_paths.append(output_path)
            output_path = output_paths[0] if output_paths else None
        
        elif config.output_format == 'json':
            output_path = os.path.join(config.output_dir, f'report_{timestamp}.json')
            json_data = {name: df.to_dict('records') for name, df in dataframes.items()}
            self.export_to_json(json_data, output_path)
        
        return output_path


def main():
    parser = argparse.ArgumentParser(description='报表导出工具')
    parser.add_argument('--host', default='localhost', help='数据库主机')
    parser.add_argument('--port', type=int, default=3306, help='数据库端口')
    parser.add_argument('--user', required=True, help='用户名')
    parser.add_argument('--password', required=True, help='密码')
    parser.add_argument('--database', required=True, help='数据库名')
    
    parser.add_argument('--type', choices=['leads', 'billing', 'combined'],
                        default='combined', help='报表类型')
    parser.add_argument('--format', choices=['excel', 'csv', 'json'],
                        default='excel', help='输出格式')
    parser.add_argument('--start-date', required=True, 
                        help='开始日期 (YYYY-MM-DD)')
    parser.add_argument('--end-date', required=True, 
                        help='结束日期 (YYYY-MM-DD)')
    parser.add_argument('--tenant-id', help='指定租户ID')
    parser.add_argument('--output-dir', default='./reports', help='输出目录')
    
    args = parser.parse_args()
    
    exporter = ReportExporter(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.database
    )
    
    if not exporter.connect():
        sys.exit(1)
    
    try:
        config = ReportConfig(
            report_type=args.type,
            start_date=datetime.strptime(args.start_date, '%Y-%m-%d'),
            end_date=datetime.strptime(args.end_date, '%Y-%m-%d'),
            output_format=args.format,
            tenant_id=args.tenant_id,
            output_dir=args.output_dir
        )
        
        output_path = exporter.generate_full_report(config)
        logger.info(f"报表生成完成: {output_path}")
    
    finally:
        exporter.close()


if __name__ == '__main__':
    main()